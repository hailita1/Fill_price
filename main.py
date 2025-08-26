import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import os

# ----------------- Main App -----------------
class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Công cụ xử lý Excel")
        self.root.geometry("920x560")
        self.root.minsize(880, 520)

        # states
        self.state_tt50 = {
            "khoa_path": None, "khoa_wb_read": None, "khoa_sheet_var": tk.StringVar(),
            "tt50_path": None, "tt50_wb_read": None, "tt50_sheet_var": tk.StringVar(),
        }
        self.state_gia = {
            "khoa_path": None, "khoa_wb_read": None, "khoa_sheet_var": tk.StringVar(),
            "gia_path": None,  "gia_wb_read": None,  "gia_sheet_var": tk.StringVar(),
        }

        # style nhẹ
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except:
            pass
        style.configure("TButton", padding=6)
        style.configure("TLabelframe", padding=8)
        style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"))
        style.configure("Header.TLabel", font=("Segoe UI", 12, "bold"))

        # Notebook tabs
        self.nb = ttk.Notebook(root)
        self.nb.pack(fill="both", expand=True, padx=12, pady=12)

        # TT50 tab
        self.tab_tt50 = ttk.Frame(self.nb)
        self.nb.add(self.tab_tt50, text="Lấy thông tin PT/TT từ TT50")
        self.build_tab_tt50(self.tab_tt50, self.state_tt50)

        # Giá tab
        self.tab_gia = ttk.Frame(self.nb)
        self.nb.add(self.tab_gia, text="Lấy giá từ cổng BHXH")
        self.build_tab_gia(self.tab_gia, self.state_gia)

    # -------- UI TT50 --------
    def build_tab_tt50(self, parent, state):
        header = ttk.Label(parent, text="Lấy thông tin PT/TT từ TT50", style="Header.TLabel")
        header.pack(anchor="w", padx=6, pady=(6, 0))

        lf = ttk.LabelFrame(parent, text="Chọn file & sheet")
        lf.pack(fill="x", padx=6, pady=10)

        ttk.Label(lf, text="File Khoa").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.tt50_khoa_entry = ttk.Entry(lf, width=56, state="readonly")
        self.tt50_khoa_entry.grid(row=0, column=1, sticky="we", padx=6, pady=6)
        ttk.Button(lf, text="Chọn", command=lambda: self.pick_file(state, "khoa", entry=self.tt50_khoa_entry)).grid(row=0, column=2, padx=6, pady=6)
        ttk.Label(lf, text="Sheet").grid(row=0, column=3, padx=(18, 6))
        self.tt50_khoa_combo = ttk.Combobox(lf, textvariable=state["khoa_sheet_var"], state="readonly", width=22)
        self.tt50_khoa_combo.grid(row=0, column=4, padx=6, pady=6)

        ttk.Label(lf, text="File TT50").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        self.tt50_file_entry = ttk.Entry(lf, width=56, state="readonly")
        self.tt50_file_entry.grid(row=1, column=1, sticky="we", padx=6, pady=6)
        ttk.Button(lf, text="Chọn", command=lambda: self.pick_file(state, "tt50", entry=self.tt50_file_entry)).grid(row=1, column=2, padx=6, pady=6)
        ttk.Label(lf, text="Sheet").grid(row=1, column=3, padx=(18, 6))
        self.tt50_file_combo = ttk.Combobox(lf, textvariable=state["tt50_sheet_var"], state="readonly", width=22)
        self.tt50_file_combo.grid(row=1, column=4, padx=6, pady=6)

        btn = ttk.Button(parent, text="Thực hiện xử lý", command=self.process_tt50)
        btn.pack(pady=10)

        self.log_tt50 = tk.Text(parent, height=10, wrap="word")
        self.log_tt50.pack(fill="both", expand=True, padx=6, pady=(0,6))

        lf.grid_columnconfigure(1, weight=1)

    # -------- UI Giá --------
    def build_tab_gia(self, parent, state):
        header = ttk.Label(parent, text="Lấy giá từ cổng BHXH", style="Header.TLabel")
        header.pack(anchor="w", padx=6, pady=(6, 0))

        lf = ttk.LabelFrame(parent, text="Chọn file & sheet")
        lf.pack(fill="x", padx=6, pady=10)

        ttk.Label(lf, text="File Khoa").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.gia_khoa_entry = ttk.Entry(lf, width=56, state="readonly")
        self.gia_khoa_entry.grid(row=0, column=1, sticky="we", padx=6, pady=6)
        ttk.Button(lf, text="Chọn", command=lambda: self.pick_file(state, "khoa", entry=self.gia_khoa_entry)).grid(row=0, column=2, padx=6, pady=6)
        ttk.Label(lf, text="Sheet").grid(row=0, column=3, padx=(18, 6))
        self.gia_khoa_combo = ttk.Combobox(lf, textvariable=state["khoa_sheet_var"], state="readonly", width=22)
        self.gia_khoa_combo.grid(row=0, column=4, padx=6, pady=6)

        ttk.Label(lf, text="File Giá").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        self.gia_file_entry = ttk.Entry(lf, width=56, state="readonly")
        self.gia_file_entry.grid(row=1, column=1, sticky="we", padx=6, pady=6)
        ttk.Button(lf, text="Chọn", command=lambda: self.pick_file(state, "gia", entry=self.gia_file_entry)).grid(row=1, column=2, padx=6, pady=6)
        ttk.Label(lf, text="Sheet").grid(row=1, column=3, padx=(18, 6))
        self.gia_file_combo = ttk.Combobox(lf, textvariable=state["gia_sheet_var"], state="readonly", width=22)
        self.gia_file_combo.grid(row=1, column=4, padx=6, pady=6)

        btn = ttk.Button(parent, text="Thực hiện xử lý", command=self.process_gia)
        btn.pack(pady=10)

        self.log_gia = tk.Text(parent, height=10, wrap="word")
        self.log_gia.pack(fill="both", expand=True, padx=6, pady=(0,6))

        lf.grid_columnconfigure(1, weight=1)

    # -------- pick file and load sheet names --------
    def pick_file(self, state, which, entry: ttk.Entry):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")])
        if not path:
            return
        try:
            wb_read = load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở file:\n{e}")
            return

        entry.config(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, path)
        entry.config(state="readonly")

        if which == "khoa":
            state["khoa_path"] = path
            state["khoa_wb_read"] = wb_read
            combo = (self.tt50_khoa_combo if state is self.state_tt50 else self.gia_khoa_combo)
            combo["values"] = wb_read.sheetnames
            state["khoa_sheet_var"].set(wb_read.sheetnames[0])
        elif which == "tt50":
            state["tt50_path"] = path
            state["tt50_wb_read"] = wb_read
            self.tt50_file_combo["values"] = wb_read.sheetnames
            state["tt50_sheet_var"].set(wb_read.sheetnames[0])
        elif which == "gia":
            state["gia_path"] = path
            state["gia_wb_read"] = wb_read
            self.gia_file_combo["values"] = wb_read.sheetnames
            state["gia_sheet_var"].set(wb_read.sheetnames[0])

    # -------- process TT50 (thêm log dòng không tìm thấy) --------
    def process_tt50(self):
        st = self.state_tt50
        if not (st["khoa_path"] and st["tt50_path"] and st["khoa_sheet_var"].get() and st["tt50_sheet_var"].get()):
            messagebox.showwarning("Thiếu dữ liệu", "Hãy chọn đủ file & sheet Khoa và TT50.")
            return

        try:
            ws_tt50 = st["tt50_wb_read"][st["tt50_sheet_var"].get()]

            label_pt = {3: "PTĐB", 4: "PT1", 5: "PT2", 6: "PT3"}
            label_tt = {7: "TTĐB", 8: "TT1", 9: "TT2", 10: "TT3"}

            mapping = {}
            for r in range(1, ws_tt50.max_row + 1):
                key_cell = ws_tt50.cell(row=r, column=2)
                if key_cell.value is None:
                    continue
                key = str(key_cell.value).strip().lower()
                found_label = None
                for c in range(3, 7):
                    val = ws_tt50.cell(row=r, column=c).value
                    if isinstance(val, str) and val.strip().lower() == "x":
                        found_label = label_pt[c]
                        break
                if not found_label:
                    for c in range(7, 11):
                        val = ws_tt50.cell(row=r, column=c).value
                        if isinstance(val, str) and val.strip().lower() == "x":
                            found_label = label_tt[c]
                            break
                if found_label:
                    mapping[key] = found_label

            wb_write = load_workbook(st["khoa_path"], data_only=False)
            ws_write = wb_write[st["khoa_sheet_var"].get()]

            hits = misses = 0
            self.log_tt50.delete("1.0", tk.END)

            for r in range(1, ws_write.max_row + 1):
                dcell = ws_write.cell(row=r, column=4)  # cột D
                fcell = ws_write.cell(row=r, column=6)  # cột F
                if dcell.value is None:
                    fcell.value = ""
                    continue
                key = str(dcell.value).strip().lower()
                if key in mapping:
                    fcell.value = mapping[key]
                    hits += 1
                else:
                    fcell.value = ""
                    misses += 1
                    self.log_tt50.insert(tk.END, f"❌ Không tìm thấy tại dòng {r} (key='{dcell.value}')\n")

            out_path = st["khoa_path"]
            wb_write.save(out_path)

            self.log_tt50.insert(tk.END, f"\n✅ Hoàn tất. Tìm được: {hits}, Không tìm được: {misses}\n")
            self.log_tt50.insert(tk.END, f"📄 Output: {out_path}\n")
            messagebox.showinfo("Thành công", f"Đã xuất file:\n{out_path}")

        except Exception as e:
            messagebox.showerror("Lỗi xử lý TT50", str(e))

    # -------- process Giá --------
    def process_gia(self):
        st = self.state_gia
        if not (st["khoa_path"] and st["gia_path"] and st["khoa_sheet_var"].get() and st["gia_sheet_var"].get()):
            messagebox.showwarning("Thiếu dữ liệu", "Hãy chọn đủ file & sheet Khoa và Giá.")
            return

        try:
            ws_gia_read = st["gia_wb_read"][st["gia_sheet_var"].get()]
            gia_map_raw = {}
            for r in range(1, ws_gia_read.max_row + 1):
                key_cell = ws_gia_read.cell(row=r, column=5)  # E
                if key_cell.value is None:
                    continue
                key = str(key_cell.value).strip()
                extra_val = ws_gia_read.cell(row=r, column=7).value  # G
                bh_val    = ws_gia_read.cell(row=r, column=9).value  # I
                dv_val    = ws_gia_read.cell(row=r, column=10).value # J
                gia_map_raw[key] = (extra_val, bh_val, dv_val)

            wb_write = load_workbook(st["khoa_path"], data_only=False)
            ws_write = wb_write[st["khoa_sheet_var"].get()]

            def format_price(val):
                if val is None:
                    return ""
                s = str(val).strip()
                if "." in s:
                    return s
                try:
                    return "{:,}".format(int(float(s))).replace(",", ".")
                except:
                    return s

            hits = misses = 0
            self.log_gia.delete("1.0", tk.END)

            for r in range(1, ws_write.max_row + 1):
                dcell = ws_write.cell(row=r, column=4)  # từ khóa
                gcell = ws_write.cell(row=r, column=7)  # Giá BH
                hcell = ws_write.cell(row=r, column=8)  # Giá DV
                icell = ws_write.cell(row=r, column=9)  # Thêm từ cột G file Giá

                if dcell.value is None:
                    gcell.value = ""
                    hcell.value = ""
                    icell.value = ""
                    continue

                key = str(dcell.value).strip()
                if key in gia_map_raw:
                    raw_extra, raw_bh, raw_dv = gia_map_raw[key]
                    gcell.value = format_price(raw_bh)
                    hcell.value = format_price(raw_dv)
                    icell.value = raw_extra if raw_extra is not None else ""
                    hits += 1
                else:
                    gcell.value = ""
                    hcell.value = ""
                    icell.value = ""
                    misses += 1
                    self.log_gia.insert(tk.END, f"❌ Không tìm thấy tại dòng {r} (key='{key}')\n")

            out_path = st["khoa_path"]
            wb_write.save(out_path)

            self.log_gia.insert(tk.END, f"\n✅ Hoàn tất. Điền dữ liệu cho {hits} dòng, không tìm thấy {misses} dòng.\n")
            self.log_gia.insert(tk.END, f"📄 Output: {out_path}\n")
            messagebox.showinfo("Thành công", f"Đã xuất file:\n{out_path}")

        except Exception as e:
            messagebox.showerror("Lỗi xử lý Giá", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()
